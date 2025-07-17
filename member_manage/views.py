# Python standard library imports
import csv
import datetime
from datetime import date, timedelta
import hashlib
import io
import json
import os
import re
import secrets
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Django imports
from django.conf import settings
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import connection
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt

# Third-party imports
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.urls import reverse
from django.views.decorators.http import require_http_methods
import mysql.connector
import openpyxl
from datetime import datetime
from django.utils.datastructures import MultiValueDictKeyError

# In-memory token store for demo (use DB or cache in production)
RESET_TOKENS = {}

# Local imports
from .db_initializer import *

def get_db_connection():
    return mysql.connector.connect(
        host=settings.DB_HOST,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        database=settings.DB_NAME
    )

def login_view(request):
    """
    Handle user login
    """
    print("Login view accessed")
    message = None
    db_initializer = DBInitializer(settings.DB_HOST, settings.DB_USER, settings.DB_PASSWORD, settings.DB_NAME)
    db_initializer.initialize()
    if request.method == 'POST':
        print("Login POST request received")
        username = request.POST.get('username')
        password = request.POST.get('password')
        # check for super user login
        if username == 'saneoeo' and password == 'saneoeo@123':
            request.session['is_admin'] = True
            request.session['is_authenticated'] = True
            return redirect('dashboard')

        print("Login attempt with username:", username,'password:', password)
        if not username or not password:
            message = "Username and password are required"
        else:
            # Get database connection
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)

            # Hash the password for comparison
            hashed_password = hashlib.sha256(password.encode()).hexdigest()

            # Check user credentials
            cursor.execute("SELECT * FROM users WHERE username = %s AND password = %s",
                           (username, hashed_password))
            user = cursor.fetchone()

            cursor.close()
            conn.close()

            if user:
                # Set session variables
                request.session['user_id'] = user['id']
                request.session['username'] = user['username']
                request.session['is_admin'] = user['is_admin']
                request.session['is_authenticated'] = True
                # insert record in login history
                insert_record_in_login_history(request,user)
                # Redirect to dashboard or home
                return redirect('dashboard')
            else:
                message = "Invalid username or password.Please try again."
    # fetch user category

    return render(request, 'login.html', {'message': message})

def insert_record_in_login_history(request, user):
    print("Inserting record in login history for user:", user['username'])
    # Ensure session_key exists
    if not request.session.session_key:
        request.session.save()
    session_key = request.session.session_key
    print("Inserting session key:", session_key)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO logged_in_users (user_id, username, session_key) VALUES (%s, %s, %s)",
        (user['id'], user['username'], session_key)
    )
    conn.commit()
    cursor.close()
    conn.close()


def list_logged_in_users(request):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT user_id, username, login_time FROM logged_in_users")
    users = cursor.fetchall()
    cursor.close()
    conn.close()
    return JsonResponse({'logged_in_users': users})
def get_events():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT id, event_name FROM event_registrations ORDER BY event_name DESC")
    events = cursor.fetchall()
    cursor.close()
    conn.close()
    return events

def home(request):
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')
    print("Home view accessed")
    instructors = get_instructors()
    events = get_events()
    message = request.session.pop('message', None)
    #fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    return render(request, 'home.html', {'message': message,'instructors': instructors,
                                         'events': events,'user_category':user_category })

@csrf_exempt
def register_member(request):
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        number = request.POST.get('number', '').strip()
        email = request.POST.get('email', '').strip()
        age= request.POST.get('age', '').strip()
        gender = request.POST.get('gender', '').strip()
        address = request.POST.get('address', '').strip()
        country= request.POST.get('member_country', '').strip()
        state = request.POST.get('member_state', '').strip()
        district = request.POST.get('member_city', '').strip()
        company = request.POST.get('company', '').strip()
        feedback = request.POST.get('feedback', '').strip()
        instructor_id = request.POST.get('instructor', '').strip()
        date_of_initiation = request.POST.get('date_of_initiation', '').strip()
        event = request.POST.get('event', '').strip()
        # fetch user category
        isAdminUser = get_user_category(request.session['username'])
        print("User category fetched admin status:", isAdminUser)
        if isAdminUser == True:
            user_category = 'admin'
        else:
            user_category = 'standard'
        # validate the form fields
        form_data = request.POST.copy()
        if not name:
            instructors = get_instructors()
            events = get_events()
            return render(request, 'home.html', {
                'message': 'Name is required.',
                'message_type': 'error',
                'form_data': form_data,
                'instructors': instructors,
                'events': events,
                'user_category': user_category
            })
        if not number or len(number) < 10 or number.isdigit() is False:
            instructors = get_instructors()
            events = get_events()
            return render(request, 'home.html', {
                'message': 'Number is missing or invalid(min 10 digits required).',
                'message_type': 'error',
                'form_data': form_data,
                'instructors': instructors,
                'events': events,
                'user_category': user_category
            })
        if email:
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_regex, email):
                instructors = get_instructors()
                events = get_events()
                return render(request, 'home.html', {
                    'message': 'Invalid email format.',
                    'message_type': 'error',
                    'form_data': form_data,
                    'instructors': instructors,
                    'events': events,
                    'user_category': user_category
                })
        if name and number and instructor_id and date_of_initiation:
            conn = mysql.connector.connect(
                host=settings.DB_HOST,
                user=settings.DB_USER,
                password=settings.DB_PASSWORD,
                database=settings.DB_NAME
            )
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO members
                (name, number, email, address,age,gender, country, state,district,company, notes, instructor_id, date_of_initiation,event_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s,%s, %s, %s)
            """, (
                name, number, email, address,age, gender, country, state,district,company, feedback, instructor_id, date_of_initiation,event
            ))
            # Insert members for the given event into event_attendance
            cur.execute("""
                INSERT INTO event_attendance (
                    event_id,
                    member_name,
                    age,
                    contact_number,
                    gender,
                    address,
                    attended_on,
                    is_new_member
                )
                SELECT
                    event_id,
                    name,
                    age,
                    number,
                    gender,
                    address,
                    date_of_initiation,
                    1
                FROM members
                WHERE event_id = %s
            """, (event,))

            # Update total_attendance in event_registrations for the given event
            cur.execute("""
                UPDATE event_registrations
                SET total_attendance = (
                    SELECT COUNT(*) FROM event_attendance WHERE event_id = %s
                )
                WHERE id = %s
            """, (event, event))
            conn.commit()
            cur.close()
            conn.close()
            # incease the attendance of selected event by 1
            if event:
                conn = mysql.connector.connect(
                    host=settings.DB_HOST,
                    user=settings.DB_USER,
                    password=settings.DB_PASSWORD,
                    database=settings.DB_NAME
                )
                cur = conn.cursor()
                cur.execute("UPDATE event_registrations SET total_attendance = total_attendance + 1 WHERE id = %s", (event,))
                conn.commit()
                cur.close()
                conn.close()
            request.session['message'] = 'Member registered successfully!'
        else:
            request.session['message'] = 'All required fields must be filled.'
    return redirect('home')


def get_instructors():
    conn = mysql.connector.connect(
        host=settings.DB_HOST,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        database=settings.DB_NAME
    )
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, name FROM instructors")
    instructors = cur.fetchall()
    cur.close()
    conn.close()
    return instructors

def add_instructor(request):
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')
    message = None
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        number = request.POST.get('number', '').strip()
        age = request.POST.get('age')
        gender = request.POST.get('gender')
        associated_since = request.POST.get('associated_since')
        updeshta_since = request.POST.get('updeshta_since')
        address = request.POST.get('address', '').strip()
        country= request.POST.get('ins_country', '').strip()
        state = request.POST.get('ins_state', '').strip()
        district = request.POST.get('ins_district', '').strip()
        is_active = request.POST.get('is_active', '1')  # Default to active


        if name and age and number and len(number) >= 10 and number.isdigit():
            conn = mysql.connector.connect(
                host=settings.DB_HOST,
                user=settings.DB_USER,
                password=settings.DB_PASSWORD,
                database=settings.DB_NAME
            )

            cur = conn.cursor()
            # Check if instructor with same number already exists
            cur.execute("SELECT id FROM instructors WHERE number = %s", (number,))
            if cur.fetchone():
                cur.close()
                conn.close()
                message = "Instructor with this contact already exists."
            else:
                cur.execute("""
                    INSERT INTO instructors (name, number, age, gender, associated_since, updeshta_since, address,state, district, country, is_active)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (name, number,age or None, gender or None, associated_since or None, updeshta_since or None, address or None,state, district, country, is_active))
                conn.commit()
                cur.close()
                conn.close()
                message = "Instructor added successfully!"
        else:
            if not name:
                message = "Name is required."
            elif not number or len(number) < 10 or not number.isdigit():
                message = "Number is missing or invalid (minimum 10 digits required)."
            elif not age:
                message = "Age is required"
            else:
                message = "Please fill all required fields correctly."
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    return render(request, 'add_instructor.html', {'message': message,'user_category': user_category})

def add_public_instructor(request):
    # check is user is authenticated
    print("Add public instructor view accessed")
    message = None
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        number = request.POST.get('number', '').strip()
        age = request.POST.get('age')
        gender = request.POST.get('gender')
        associated_since = request.POST.get('associated_since')
        updeshta_since = request.POST.get('updeshta_since')
        address = request.POST.get('address', '').strip()
        country= request.POST.get('ins_country', '').strip()
        state = request.POST.get('ins_state', '').strip()
        district = request.POST.get('ins_district', '').strip()

        is_active = request.POST.get('is_active', '1')  # Default to active
        if name:
            conn = mysql.connector.connect(
                host=settings.DB_HOST,
                user=settings.DB_USER,
                password=settings.DB_PASSWORD,
                database=settings.DB_NAME
            )
            cur = conn.cursor()
            cur.execute("""
                INSERT INTO instructors (name, number, age,gender, associated_since, updeshta_since, address,state, district, country, is_active)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (name, number, age or None, gender or None, associated_since or None, updeshta_since or None, address or None,state, district, country, is_active))
            conn.commit()
            cur.close()
            conn.close()
            return redirect('thank_you')
        else:
            message = "Name is required."

    return render(request, 'public_instructor_register.html', {'message': message})

def all_members(request):
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')
    print("All members view accessed")
    # Get database connection
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get all unique companies for filter dropdown
    cursor.execute("SELECT DISTINCT company FROM members WHERE company IS NOT NULL AND company != ''")
    companies = [row['company'] for row in cursor.fetchall()]

    # Get all instructors for filter dropdown
    cursor.execute("SELECT id, name FROM instructors")
    instructors = cursor.fetchall()

    # Get all events for filter dropdown
    cursor.execute("SELECT id, event_name FROM event_registrations")
    events = cursor.fetchall()

    # Fetch all members with their instructor names and event name
    cursor.execute("""
        SELECT
            m.id, m.name, m.number, m.email, m.age, m.gender, m.address,
            city.name AS district,
            state.name AS state,
            country.name AS country,
            m.company, m.instructor_id, m.date_of_initiation, m.notes,
            i.name as instructor_name,
            e.event_name as event_name
        FROM
            members m
        LEFT JOIN
            instructors i ON m.instructor_id = i.id
        LEFT JOIN
            country ON m.country = country.id
        LEFT JOIN
            state ON m.state = state.id
        LEFT JOIN
            city ON m.district = city.id
        LEFT JOIN
            event_registrations e ON m.event_id = e.id
        ORDER BY m.name
    """)
    members = cursor.fetchall()
    print("Members fetched:", len(members))

    # Format dates for display
    for member in members:
        if member['date_of_initiation']:
            member['date_of_initiation'] = member['date_of_initiation'].strftime('%Y-%m-%d')

    cursor.close()
    conn.close()
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    context = {
        'companies': companies,
        'instructors': instructors,
        'events': events,
        'members': members,  # This data will be displayed in the template
        'member_count': len(members),  # Total number of members
        'user_category': user_category
    }

    return render(request, 'all_members.html', context)

@require_http_methods(["GET", "DELETE"])
def api_member_detail(request, id):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "DELETE":
        try:
            cursor.execute("DELETE FROM members WHERE id = %s", (id,))
            conn.commit()
            affected_rows = cursor.rowcount
            cursor.close()
            conn.close()

            if affected_rows > 0:
                return JsonResponse({'success': True, 'message': 'Member deleted successfully'})
            else:
                return JsonResponse({'error': 'Member not found'}, status=404)

        except Exception as e:
            conn.rollback()
            cursor.close()
            conn.close()
            return JsonResponse({'error': str(e)}, status=500)

    # GET request
    cursor.execute("""
        SELECT
            m.id, m.name, m.number, m.email,m.age,m.gender, m.address, m.company,
            m.instructor_id, m.date_of_initiation, m.notes,
            i.name as instructor_name
        FROM
            members m
        LEFT JOIN
            instructors i ON m.instructor_id = i.id
        WHERE
            m.id = %s
    """, (id,))

    member = cursor.fetchone()
    cursor.close()
    conn.close()

    if not member:
        return JsonResponse({'error': 'Member not found'}, status=404)

    # Convert date to string format for JSON serialization
    if member['date_of_initiation']:
        member['date_of_initiation'] = member['date_of_initiation'].isoformat()

    member_data = {
        'id': member['id'],
        'name': member['name'],
        'number': member['number'],
        'email': member['email'],
        'age': member['age'],
        'gender': member['gender'],
        'address': member['address'],
        'company': member['company'],
        'instructor_id': member['instructor_id'],
        'instructor_name': member['instructor_name'] if member['instructor_name'] else 'N/A',
        'date_of_initiation': member['date_of_initiation'],
        'notes': member['notes']
    }

    return JsonResponse(member_data)

def api_members(request):
    print("API members view accessed")
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            m.id, m.name, m.number, m.email, m.address, m.company,
            m.instructor_id, m.date_of_initiation, m.notes,
            i.name as instructor_name
        FROM
            members m
        LEFT JOIN
            instructors i ON m.instructor_id = i.id
    """)

    members = cursor.fetchall()
    cursor.close()
    conn.close()

    # Convert to list of dictionaries for JSON response
    members_list = []
    for member in members:
        # Convert date to string format for JSON serialization
        if member['date_of_initiation']:
            member['date_of_initiation'] = member['date_of_initiation'].isoformat()

        members_list.append({
            'id': member['id'],
            'name': member['name'],
            'number': member['number'],
            'email': member['email'],
            'address': member['address'],
            'company': member['company'],
            'instructor_id': member['instructor_id'],
            'instructor_name': member['instructor_name'] if member['instructor_name'] else 'N/A',
            'date_of_initiation': member['date_of_initiation'],
            'notes': member['notes']
        })

    return JsonResponse(members_list, safe=False)

def all_instructors(request):
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')
    print("All instructors view accessed")
    success_message = request.session.pop('success_message', None)

    # Get database connection
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Get all instructors with count of members they teach

    cursor.execute("""
        SELECT
            i.id, i.name, i.number, i.age, i.gender, i.associated_since,
            i.updeshta_since, i.address,
            d.name AS district,
            s.name AS state,
            c.name AS country,
            i.is_active,
            COUNT(m.id) as member_count
        FROM
            instructors i
        LEFT JOIN
            members m ON i.id = m.instructor_id
        LEFT JOIN
            country c ON i.country = c.id
        LEFT JOIN
            state s ON i.state = s.id
        LEFT JOIN
            city d ON i.district = d.id
        GROUP BY
            i.id
        ORDER BY
            i.name
    """)
    instructors = cursor.fetchall()
    print(f"Instructors fetched: {len(instructors)}")

    # Format dates for display
    for instructor in instructors:
        # Get associated_since and updeshta_since years as strings
        instructor['associated_since'] = str(instructor['associated_since']) if instructor['associated_since'] else 'N/A'
        instructor['updeshta_since'] = str(instructor['updeshta_since']) if instructor['updeshta_since'] else 'N/A'

    # Get distinct years for filters
    cursor.execute("SELECT DISTINCT associated_since FROM instructors WHERE associated_since IS NOT NULL ORDER BY associated_since DESC")
    associated_years = [str(row['associated_since']) for row in cursor.fetchall()]

    cursor.execute("SELECT DISTINCT updeshta_since FROM instructors WHERE updeshta_since IS NOT NULL ORDER BY updeshta_since DESC")
    updeshta_years = [str(row['updeshta_since']) for row in cursor.fetchall()]

    # Generate range of years from 1900 to current year
    import datetime
    current_year = datetime.datetime.now().year
    range_years = list(range(current_year, 1899, -1))

    cursor.close()
    conn.close()
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    context = {
        'instructors': instructors,
        'associated_years': associated_years,
        'updeshta_years': updeshta_years,
        'range_years': range_years,
        'success_message': success_message,
        'user_category': user_category
    }
    return render(request, 'all_instructors.html', context)

def register_user(request):
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    if request.method == 'POST':
        # Generate years from 1900 to current year
        import datetime
        current_year = datetime.datetime.now().year
        range_years = list(range(current_year, 1899, -1))

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            # Handle AJAX submission
            try:
                # Get form data
                name = request.POST.get('name')
                username = request.POST.get('username')
                email = request.POST.get('email')
                age = request.POST.get('age')
                dob = request.POST.get('dob')
                associated_since = request.POST.get('associated_since')
                updeshta_since = request.POST.get('updeshta_since')
                address = request.POST.get('address')
                reason = request.POST.get('reason')
                gmail_user = 'sant.vihangam@gmail.com'
                gmail_app_password = 'pdsexaeusfdgvqsu'
                # Send email to admin
                from django.core.mail import send_mail

                subject = f"New User Registration Request: {name}"
                message = f"""
                New user registration request details:
                
                Name: {name}
                Username: {username}
                Email: {email}
                Age: {age}
                Date of Birth: {dob}
                Associated Since: {associated_since}
                Updeshta Since: {updeshta_since}
                Address: {address}
                
                Reason for Registration:
                {reason}
                """
                msg = MIMEText(message)
                msg['Subject'] = subject
                msg['From'] = "sant.vihangam@gmail.com"
                msg['To'] = "sant.vihangam@gmail.com"

                # Send email via Gmail SMTP
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login(gmail_user, gmail_app_password)
                server.sendmail("sant.vihangam@gmail.com", "sant.vihangam@gmail.com", msg.as_string())
                server.quit()

                return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            # Handle regular form submission (fallback)
            # Process form and send email
            # Redirect to success page
            return render(request, 'register_user.html', {'success': True, 'range_years': range_years
                                                          ,'user_category': user_category})
    else:
        # Generate years from 1900 to current year
        import datetime
        current_year = datetime.datetime.now().year
        range_years = list(range(current_year, 1899, -1))

        return render(request, 'register_user.html', {'range_years': range_years,
                                                      'user_category': user_category})


def dictfetchall(cursor):
    """Convert database cursor results to a list of dictionaries"""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]

def instructor_infographics(request):
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')
    """View for instructor performance infographics page"""
    # Get database connection
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    print("Instructor infographics view accessed")

    # Safe function to convert MySQL results to integers
    def safe_int(value, default=0):
        try:
            if value is None:
                return default
            return int(value)
        except (ValueError, TypeError):
            return default

    # Get all active instructors
    cursor.execute("SELECT * FROM instructors WHERE is_active = 1 ORDER BY name")
    instructors = cursor.fetchall()  # Using direct fetchall since cursor is dictionary=True

    # Get oldest initiation date
    cursor.execute("SELECT MIN(date_of_initiation) as oldest_date FROM members")
    oldest_date_result = cursor.fetchall()

    # Set start year based on oldest member
    if oldest_date_result and oldest_date_result[0].get('oldest_date'):
        oldest_date_val = oldest_date_result[0]['oldest_date']
        # Fix the isinstance check for date objects
        # Import datetime correctly at the top of your file


        # Then in your instructor_infographics function, replace the problematic line with:
        if isinstance(oldest_date_val, (datetime, date)):
            start_year = oldest_date_val.year
        elif isinstance(oldest_date_val, str):
            try:
                start_year = datetime.strptime(oldest_date_val, '%Y-%m-%d').year
            except ValueError:
                start_year = datetime.now().year - 5
        else:
            start_year = datetime.now().year - 5
    else:
        start_year = datetime.now().year - 5

    current_year = datetime.now().year
    years_range = range(start_year, current_year + 1)

    # Get total members count
    cursor.execute("SELECT COUNT(*) as total FROM members")
    total_members_result = cursor.fetchall()
    total_members = safe_int(total_members_result[0].get('total')) if total_members_result else 0

    # Get active instructors count
    cursor.execute("SELECT COUNT(*) as total FROM instructors WHERE is_active = 1")
    instructors_result = cursor.fetchall()
    instructors_count = safe_int(instructors_result[0].get('total')) if instructors_result else 0

    # Calculate average members per instructor
    avg_members_per_instructor = round(total_members / instructors_count, 1) if instructors_count > 0 else 0

    # Get current year members
    cursor.execute(
        "SELECT COUNT(*) as count FROM members WHERE YEAR(date_of_initiation) = %s",
        [current_year]
    )
    current_year_result = cursor.fetchall()
    current_year_members = safe_int(current_year_result[0].get('count')) if current_year_result else 0

    # Get previous year members
    cursor.execute(
        "SELECT COUNT(*) as count FROM members WHERE YEAR(date_of_initiation) = %s",
        [current_year-1]
    )
    previous_year_result = cursor.fetchall()
    previous_year_members = safe_int(previous_year_result[0].get('count')) if previous_year_result else 0

    # Calculate growth rate
    if previous_year_members > 0:
        growth_rate = round(((current_year_members - previous_year_members) / previous_year_members * 100), 1)
    else:
        growth_rate = 0

    # Close database connection
    cursor.close()
    conn.close()
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    context = {
        'instructors': instructors,
        'years_range': years_range,
        'total_members': total_members,
        'instructors_count': instructors_count,
        'avg_members_per_instructor': avg_members_per_instructor,
        'growth_rate': growth_rate,
        'user_category': user_category
    }

    return render(request, 'instructor_infographics.html', context)

def api_instructor_infographics_data(request):
    """API endpoint for instructor infographics chart data"""
    # Get filter parameters
    instructor_id = request.GET.get('instructor_id', 'all')
    time_range = request.GET.get('time_range', 'all')
    year = request.GET.get('year', 'all')

    # Apply filters to member query
    member_filter_params = []
    member_filter_sql = ""

    if instructor_id != 'all':
        member_filter_sql += " AND m.instructor_id = %s"
        member_filter_params.append(instructor_id)

    # Apply time range filter
    current_date = datetime.now()
    if time_range == 'year':
        start_date = datetime(current_date.year, 1, 1)
        member_filter_sql += " AND m.date_of_initiation >= %s"
        member_filter_params.append(start_date)
    elif time_range == 'quarter':
        start_date = current_date - timedelta(days=90)
        member_filter_sql += " AND m.date_of_initiation >= %s"
        member_filter_params.append(start_date)
    elif time_range == 'month':
        start_date = current_date - timedelta(days=30)
        member_filter_sql += " AND m.date_of_initiation >= %s"
        member_filter_params.append(start_date)

    if year != 'all':
        member_filter_sql += " AND YEAR(m.date_of_initiation) = %s"
        member_filter_params.append(year)

    # Get database connection
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Get active instructors
        cursor.execute("SELECT id, name FROM instructors WHERE is_active = 1")
        instructors = cursor.fetchall()

        # Prepare data for Members by Instructor chart
        members_by_instructor_labels = []
        members_by_instructor_data = []

        for instructor in instructors:
            query = """
                SELECT COUNT(*) as member_count 
                FROM members m
                WHERE m.instructor_id = %s
            """ + member_filter_sql

            params = [instructor['id']] + member_filter_params
            cursor.execute(query, params)
            result = cursor.fetchone()
            if result and result['member_count'] > 0:
                members_by_instructor_labels.append(instructor['name'])
                members_by_instructor_data.append(result['member_count'])

        # Prepare data for Member Growth Over Time chart
        growth_labels = []
        growth_data = []

        if time_range == 'month':
            # For month view, group by day
            for i in range(30):
                day = current_date - timedelta(days=i)
                next_day = day + timedelta(days=1)

                query = """
                    SELECT COUNT(*) as count 
                    FROM members m
                    WHERE m.date_of_initiation >= %s 
                    AND m.date_of_initiation < %s
                """ + member_filter_sql

                params = [day.replace(hour=0, minute=0, second=0),
                          next_day.replace(hour=0, minute=0, second=0)] + member_filter_params

                cursor.execute(query, params)
                result = cursor.fetchone()
                count = result['count'] if result else 0
                growth_labels.insert(0, day.strftime('%b %d'))
                growth_data.insert(0, count)
        else:
            # Default view by month for the last 12 months
            for i in range(48):
                month_start = (current_date.replace(day=1) - timedelta(days=i * 30))
                next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)

                query = """
                    SELECT COUNT(*) as count 
                    FROM members m
                    WHERE m.date_of_initiation >= %s 
                    AND m.date_of_initiation < %s
                """ + member_filter_sql

                params = [month_start, next_month] + member_filter_params

                cursor.execute(query, params)
                result = cursor.fetchone()
                count = result['count'] if result else 0
                growth_labels.insert(0, month_start.strftime('%b %Y'))
                growth_data.insert(0, count)

        # Prepare data for Geographic Distribution chart
        # Prepare data for Geographic Distribution chart
        if instructor_id != 'all':
            # city-wise distribution for selected instructor
            query = """
                SELECT district as location, COUNT(*) as count 
                FROM members m
                WHERE district IS NOT NULL AND district != ''
            """ + member_filter_sql + """
                GROUP BY district 
                ORDER BY count DESC
                LIMIT 8
            """
        else:
            # country-wise distribution for all instructors
            query = """
                SELECT country as location, COUNT(*) as count 
                FROM members m
                WHERE country IS NOT NULL AND country != ''
            """ + member_filter_sql + """
                GROUP BY country 
                ORDER BY count DESC
                LIMIT 8
            """

        cursor.execute(query, member_filter_params)
        geo_distribution = cursor.fetchall()

        geo_labels = []
        geo_data = []

        for entry in geo_distribution:
            if entry.get('location'):
                geo_labels.append(entry.get('location'))
                geo_data.append(entry.get('count'))

        # If no locations found, add placeholder
        if not geo_labels:
            geo_labels = ['No Data']
            geo_data = [0]

        # Prepare data for Instructor Comparison chart
        # Get top 5 instructors by member count
        top_instructors_data = []

        for instructor in instructors:
            query = """
                SELECT COUNT(*) as member_count 
                FROM members m
                WHERE m.instructor_id = %s
            """ + member_filter_sql

            params = [instructor['id']] + member_filter_params
            cursor.execute(query, params)
            result = cursor.fetchone()
            member_count = result['member_count'] if result else 0

            if member_count > 0:
                top_instructors_data.append({
                    'id': instructor['id'],
                    'name': instructor['name'],
                    'member_count': member_count
                })

        # Sort and limit to top 5
        top_instructors_data = sorted(top_instructors_data,
                                      key=lambda x: x['member_count'],
                                      reverse=True)[:5]

        comparison_datasets = []
        colors = [
            'rgba(79, 140, 255, 0.7)',
            'rgba(110, 214, 255, 0.7)',
            'rgba(255, 159, 64, 0.7)',
            'rgba(75, 192, 192, 0.7)',
            'rgba(255, 99, 132, 0.7)'
        ]

        for i, instructor in enumerate(top_instructors_data):
            # Get additional data about instructor
            cursor.execute("SELECT * FROM instructors WHERE id = %s", [instructor['id']])
            instructor_details = cursor.fetchone()

            # Calculate metrics (scaled 0-100 for radar chart)
            member_count_score = min(100, instructor['member_count'] * 5)  # Assuming 20 members = 100%

            # Years of experience score (max 10 years = 100%)
            associated_since = instructor_details.get('associated_since') if instructor_details else None
            experience_years = (datetime.now().year - associated_since) if associated_since else 0
            experience_score = min(100, experience_years * 10)

            # Other metrics are placeholders - would need more data in real app
            growth_rate = 60 + i * 5
            retention_rate = 75 - i * 3
            activity_score = 70 + i * 4

            comparison_datasets.append({
                'label': instructor['name'],
                'data': [member_count_score, growth_rate, retention_rate, activity_score, experience_score],
                'backgroundColor': colors[i % len(colors)],
                'borderColor': colors[i % len(colors)].replace('0.7', '1'),
                'pointBackgroundColor': colors[i % len(colors)].replace('0.7', '1'),
                'pointBorderColor': '#fff',
                'pointHoverBackgroundColor': '#fff',
                'pointHoverBorderColor': colors[i % len(colors)].replace('0.7', '1')
            })

        # Calculate summary metrics
        query = "SELECT COUNT(*) as total FROM members m WHERE 1=1" + member_filter_sql
        cursor.execute(query, member_filter_params)
        total_members = cursor.fetchone()['total']

        query = "SELECT COUNT(*) as total FROM instructors WHERE is_active = 1"
        cursor.execute(query)
        active_instructors = cursor.fetchone()['total']

        avg_members_per_instructor = round(total_members / active_instructors, 1) if active_instructors > 0 else 0

        # Calculate growth rate
        current_year = datetime.now().year

        cursor.execute(
            "SELECT COUNT(*) as count FROM members m WHERE YEAR(m.date_of_initiation) = %s" + member_filter_sql,
            [current_year] + member_filter_params
        )
        current_year_members = cursor.fetchone()['count']

        cursor.execute(
            "SELECT COUNT(*) as count FROM members m WHERE YEAR(m.date_of_initiation) = %s" + member_filter_sql,
            [current_year - 1] + member_filter_params
        )
        previous_year_members = cursor.fetchone()['count']

        growth_rate = round(((current_year_members - previous_year_members) / max(1, previous_year_members) * 100), 1)

        # Prepare response data
        response_data = {
            'membersByInstructor': {
                'labels': members_by_instructor_labels,
                'data': members_by_instructor_data
            },
            'memberGrowth': {
                'labels': growth_labels,
                'data': growth_data
            },
            'geoDistribution': {
                'labels': geo_labels,
                'data': geo_data
            },
            'instructorComparison': {
                'datasets': comparison_datasets
            },
            'summary': {
                'totalMembers': total_members,
                'activeInstructors': active_instructors,
                'avgMembersPerInstructor': avg_members_per_instructor,
                'growthRate': growth_rate
            }
        }

        return JsonResponse(response_data)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        cursor.close()
        conn.close()

def api_instructor_details(request, instructor_id):
    from datetime import datetime
    """API endpoint for detailed information about a specific instructor"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    print("API instructor details view accessed")
    try:
        # Get instructor details
        cursor.execute("SELECT * FROM instructors WHERE id = %s", [instructor_id])
        instructor = cursor.fetchone()

        if not instructor:
            return JsonResponse({'error': 'Instructor not found'}, status=404)

        # Count total members for this instructor
        cursor.execute(
            "SELECT COUNT(*) as total FROM members WHERE instructor_id = %s",
            [instructor_id]
        )
        total_members = cursor.fetchone()['total']

        # Calculate growth rate (members joined this year vs last year)
        current_year = datetime.now().year
        cursor.execute(
            "SELECT COUNT(*) as count FROM members WHERE instructor_id = %s AND YEAR(date_of_initiation) = %s",
            [instructor_id, current_year]
        )
        current_year_members = cursor.fetchone()['count']

        cursor.execute(
            "SELECT COUNT(*) as count FROM members WHERE instructor_id = %s AND YEAR(date_of_initiation) = %s",
            [instructor_id, current_year - 1]
        )
        previous_year_members = cursor.fetchone()['count']

        if previous_year_members > 0:
            growth_rate = round(((current_year_members - previous_year_members) / previous_year_members) * 100, 1)
        else:
            growth_rate = 0

        # Calculate retention rate (placeholder logic)
        cursor.execute(
            "SELECT COUNT(*) as count FROM members WHERE instructor_id = %s AND date_of_initiation <= %s",
            [instructor_id, f"{current_year-1}-12-31"]
        )
        retained_members = cursor.fetchone()['count']
        retention_rate = round((retained_members / total_members) * 100, 1) if total_members > 0 else 0

        # Member growth trend (last 12 months)
        trend_labels = []
        trend_data = []
        for i in range(12):
            month = (datetime.now().replace(day=1) - timedelta(days=i*30))
            next_month = (month.replace(day=28) + timedelta(days=4)).replace(day=1)
            cursor.execute(
                "SELECT COUNT(*) as count FROM members WHERE instructor_id = %s AND date_of_initiation >= %s AND date_of_initiation < %s",
                [instructor_id, month, next_month]
            )
            count = cursor.fetchone()['count']
            trend_labels.insert(0, month.strftime('%b %Y'))
            trend_data.insert(0, count)

        data = {
            'id': instructor['id'],
            'name': instructor['name'],
            'associated_since': str(instructor['associated_since']) if instructor['associated_since'] else 'N/A',
            'total_members': total_members,
            'growth_rate': growth_rate,
            'retention_rate': retention_rate,
            'trend': {
                'labels': trend_labels,
                'data': trend_data
            }
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    finally:
        cursor.close()
        conn.close()

def api_download_instructor_report(request):
    """API endpoint to download instructor report as CSV"""
    # Get filter parameters
    instructor_id = request.GET.get('instructor_id', 'all')
    time_range = request.GET.get('time_range', 'all')
    year = request.GET.get('year', 'all')

    # Apply filters to member query
    member_filter_params = []
    member_filter_sql = ""

    if instructor_id != 'all':
        member_filter_sql += " AND instructor_id = %s"
        member_filter_params.append(instructor_id)

    # Apply time range filter
    current_date = datetime.now()
    if time_range == 'year':
        start_date = datetime(current_date.year, 1, 1)
        member_filter_sql += " AND date_of_initiation >= %s"
        member_filter_params.append(start_date)
    elif time_range == 'quarter':
        start_date = current_date - timedelta(days=90)
        member_filter_sql += " AND date_of_initiation >= %s"
        member_filter_params.append(start_date)
    elif time_range == 'month':
        start_date = current_date - timedelta(days=30)
        member_filter_sql += " AND date_of_initiation >= %s"
        member_filter_params.append(start_date)

    if year != 'all':
        member_filter_sql += " AND YEAR(date_of_initiation) = %s"
        member_filter_params.append(year)

    # Create CSV file
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Write header row
    writer.writerow(['Instructor Name', 'Total Members', 'New Members This Year', 'Countries', 'states'])

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    try:
        # Get data for all instructors or the selected one
        if instructor_id != 'all':
            cursor.execute("SELECT * FROM instructors WHERE is_active = 1 AND id = %s", [instructor_id])
        else:
            cursor.execute("SELECT * FROM instructors WHERE is_active = 1")

        instructors = cursor.fetchall()

        for instructor in instructors:
            # Get total members for this instructor
            cursor.execute(
                "SELECT COUNT(*) as total FROM members WHERE instructor_id = %s" + member_filter_sql,
                [instructor['id']] + member_filter_params
            )
            total_members = cursor.fetchone()['total']

            # Get new members this year
            cursor.execute(
                """
                SELECT COUNT(*) as count 
                FROM members 
                WHERE instructor_id = %s AND YEAR(date_of_initiation) = %s
                """ + member_filter_sql,
                [instructor['id'], current_date.year] + member_filter_params
            )
            new_members_this_year = cursor.fetchone()['count']

            # Get top countries
            cursor.execute(
                """
                SELECT country, COUNT(*) as count 
                FROM members 
                WHERE instructor_id = %s AND country IS NOT NULL AND country != ''
                """ + member_filter_sql + """
                GROUP BY country 
                ORDER BY count DESC 
                LIMIT 3
                """,
                [instructor['id']] + member_filter_params
            )
            countries = cursor.fetchall()
            countries_str = ", ".join([f"{c['country']} ({c['count']})" for c in countries if c['country']])

            # Get top states
            cursor.execute(
                """
                SELECT state, COUNT(*) as count 
                FROM members 
                WHERE instructor_id = %s AND state IS NOT NULL AND state != ''
                """ + member_filter_sql + """
                GROUP BY state 
                ORDER BY count DESC 
                LIMIT 3
                """,
                [instructor['id']] + member_filter_params
            )
            states = cursor.fetchall()
            states_str = ", ".join([f"{s['state']} ({s['count']})" for s in states if s['state']])

            # Write data row
            writer.writerow([
                instructor['name'],
                total_members,
                new_members_this_year,
                countries_str,
                states_str
            ])

    finally:
        cursor.close()
        conn.close()

    # Prepare the CSV file for download
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=instructor_report_{datetime.now().strftime("%Y%m%d")}.csv'

    return response

@csrf_exempt
def upload_members(request):
    # check is user is authenticated
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    if not request.session.get('is_authenticated'):
        return redirect('login')
    if request.method == 'GET':
        return render(request, 'upload_members.html',{'user_category': user_category})
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            required = ['name', 'number', 'instructor_id', 'date_of_initiation']
            db_columns = ['name', 'number', 'email', 'age','gender', 'address', 'state', 'district', 'country', 'company', 'notes', 'instructor_id', 'event_id','date_of_initiation']
            # Validate columns
            if not all(col in headers for col in required):
                return JsonResponse({'success': False, 'message': 'Missing required columns in Excel file.'})
            # Prepare data rows
            rows = []
            event_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if not all(row_dict.get(col) for col in required):
                    continue  # skip incomplete rows
                rows.append([row_dict.get(col, None) for col in db_columns])
                if 'event_id' in headers and row_dict.get('event_id'):
                    event_ids.add(row_dict.get('event_id'))
            if not rows:
                return JsonResponse({'success': False, 'message': 'No valid rows to import.'})
            # Insert into DB
            import mysql.connector
            from django.conf import settings
            conn = mysql.connector.connect(
                host=settings.DB_HOST,
                user=settings.DB_USER,
                password=settings.DB_PASSWORD,
                database=settings.DB_NAME
            )
            cur = conn.cursor()
            sql = f"""INSERT INTO members
                (name, number, email,age,gender, address, state, district, country, company, notes, instructor_id,event_id, date_of_initiation)
                VALUES ({','.join(['%s']*14)})"""
            cur.executemany(sql, rows)
            # For each event_id, insert into event_attendance and update event_registrations
            for event in event_ids:
                cur.execute("""
                    INSERT INTO event_attendance (
                        event_id,
                        member_name,
                        age,
                        contact_number,
                        gender,
                        address,
                        attended_on,
                        is_new_member
                    )
                    SELECT
                        event_id,
                        name,
                        age,
                        number,
                        gender,
                        address,
                        date_of_initiation,
                        1
                    FROM members
                    WHERE event_id = %s
                """, (event,))
                cur.execute("""
                    UPDATE event_registrations
                    SET total_attendance = (
                        SELECT COUNT(*) FROM event_attendance WHERE event_id = %s
                    )
                    WHERE id = %s
                """, (event, event))
            conn.commit()
            cur.close()
            conn.close()
            return JsonResponse({'success': True, 'message': f'Successfully imported {len(rows)} members.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Invalid request.'})

def change_password(request):
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')
    """View for changing user passwords"""
    # Get a database connection
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    # Fetch all users for the dropdown
    cursor.execute("SELECT id, username FROM users")
    users = cursor.fetchall()

    message = None
    success = False

    if request.method == 'POST':
        user_id = request.POST.get('user')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        # Validate inputs
        if not all([user_id, new_password, confirm_password]):
            message = "All fields are required"
        elif new_password != confirm_password:
            message = "New passwords don't match"
        else:
            # Validate password strength
            if not validate_password_strength(new_password):
                message = "Password must be at least 8 characters and include letters, numbers, and a special character"
            else:
                # Update with new password
                hashed_new = hashlib.sha256(new_password.encode()).hexdigest()
                cursor.execute("UPDATE users SET password = %s WHERE id = %s", (hashed_new, user_id))
                conn.commit()
                message = "Password changed successfully"
                success = True


    cursor.close()
    conn.close()
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    return render(request, 'password_change.html', {
        'users': users,
        'message': message,
        'success': success,
        'user_category': user_category
    })

def validate_password_strength(password):
    """Validate password meets complexity requirements"""
    if len(password) < 8:
        return False

    # Check for alphanumeric + special char
    has_letter = bool(re.search(r'[a-zA-Z]', password))
    has_number = bool(re.search(r'\d', password))
    has_special = bool(re.search(r'[!@#$%^&*()_+\-=\[\]{};\':"\\|,.<>/?]', password))

    return has_letter and has_number and has_special


def logout_view(request):
    """
    Logs out the user and terminates the session
    """
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')
    # Clear session variables
    if 'user_id' in request.session:
        user_id = request.session['user_id']
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM logged_in_users WHERE user_id = %s", (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
        del request.session['user_id']

    if 'is_authenticated' in request.session:
        del request.session['is_authenticated']
    if 'is_admin' in request.session:
        del request.session['is_admin']

    if 'username' in request.session:
        del request.session['username']

    # Flush the session
    request.session.flush()

    # Render the logout page
    return render(request, 'logout.html')

# views.py

def get_member_detail(request, member_id):
    print("API member detail view accessed")
    """API endpoint to get member details using raw SQL"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Get member data with country, state, district (city) names, instructor, and event name
        cursor.execute("""
            SELECT
                m.*,
                city.name AS district,
                state.name AS state,
                country.name AS country,
                i.name as instructor_name,
                e.event_name as event_name
            FROM
                members m
            LEFT JOIN
                instructors i ON m.instructor_id = i.id
            LEFT JOIN
                country ON m.country = country.id
            LEFT JOIN
                state ON m.state = state.id
            LEFT JOIN
                city ON m.district = city.id
            LEFT JOIN
                event_registrations e ON m.event_id = e.id
            WHERE
                m.id = %s
        """, (member_id,))

        member = cursor.fetchone()

        if not member:
            return JsonResponse({'error': 'Member not found'}, status=404)

        # Get all active instructors
        cursor.execute("""
            SELECT id, name
            FROM instructors
            WHERE is_active = 1
        """)
        instructors = cursor.fetchall()

        # Format date for JSON serialization
        if member['date_of_initiation']:
            member['date_of_initiation'] = member['date_of_initiation'].strftime('%Y-%m-%d')

        # Prepare the response data
        member_data = {
            'id': member['id'],
            'name': member['name'],
            'number': member['number'],
            'email': member['email'],
            'age': member['age'],
            'gender': member['gender'],
            'address': member['address'],
            'state': member['state'],
            'district': member['district'],
            'country': member['country'],
            'company': member['company'],
            'notes': member['notes'],
            'instructor_id': member['instructor_id'],
            'date_of_initiation': member['date_of_initiation'],
            'event_name': member['event_name'],
            'instructors': [{'id': i['id'], 'name': i['name']} for i in instructors]
        }

        cursor.close()
        conn.close()

        return JsonResponse(member_data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@require_http_methods(["POST"])
def update_member(request, member_id):
    print("API update member view accessed")
    """API endpoint to update member details using raw SQL"""
    try:
        data = json.loads(request.body)
        conn = get_db_connection()
        cursor = conn.cursor()
        print("Data received for update:", data)
        # Format instructor_id - set to NULL if empty
        instructor_id = data.get('instructor_id')
        if not instructor_id or instructor_id.strip() == '':
            instructor_id = None

        # Correct usage if 'from datetime import datetime' is present
        date_of_initiation = data.get('date_of_initiation')
        if date_of_initiation:
            date_of_initiation = datetime.strptime(date_of_initiation, '%Y-%m-%d').date()

        # Update member record
        cursor.execute("""
            UPDATE members
            SET name = %s, 
                number = %s, 
                email = %s, 
                age = %s, 
                gender = %s, 
                address = %s, 
                state = %s, 
                district = %s, 
                country = %s, 
                company = %s, 
                notes = %s, 
                instructor_id = %s, 
                date_of_initiation = %s
            WHERE id = %s
        """, (
            data.get('name'),
            data.get('number'),
            data.get('email'),
            data.get('age'),
            data.get('gender'),
            data.get('address'),
            data.get('state'),
            data.get('district'),
            data.get('country'),
            data.get('company'),
            data.get('notes'),
            instructor_id,
            date_of_initiation,
            member_id
        ))

        conn.commit()
        cursor.close()
        conn.close()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def delete_member(request, member_id):
    """API endpoint to delete a member using raw SQL"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # Delete member record
        cursor.execute("DELETE FROM members WHERE id = %s", (member_id,))

        conn.commit()
        cursor.close()
        conn.close()

        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

def username_exists(username):
    """
    Check if a username already exists in the users table.
    Returns True if exists, False otherwise.
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT 1 FROM users WHERE username = %s LIMIT 1", (username,))
    exists = cursor.fetchone() is not None
    cursor.close()
    conn.close()
    return exists

def create_user(request):
    """View for creating a new user with admin privileges"""
    if not request.session.get('is_authenticated'):
        return redirect('login')

    message = None
    success = False
    print("Create user view accessed")

    conn = get_db_connection()
    cursor = conn.cursor()

    # Fetch all users for the table
    cursor.execute("""
        SELECT id, username, email, is_admin, created_on
        FROM users
        ORDER BY created_on DESC
    """)
    all_users = cursor.fetchall()

    # Set up pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(all_users, 5)
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    if request.method == 'POST':
        username = request.POST.get('username')
        name = request.POST.get('name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        user_category = request.POST.get('user_category')

        if username_exists(username):
            print("Username already exists:", username)
            message = f"Username {username} already exists!"
            success = False
        elif password != confirm_password:
            message = "Passwords do not match!"
            success = False
        elif not validate_password_strength(password):
            message = "Password does not meet complexity requirements!"
            success = False
        else:
            try:
                hashed_password = hashlib.sha256(password.encode()).hexdigest()
                is_admin = 1 if user_category == 'Admin' else 0
                cursor.execute("""
                    INSERT INTO users (username, password, email, is_admin)
                    VALUES (%s, %s, %s, %s)
                """, (username, hashed_password, email, is_admin))
                conn.commit()
                message = f"User {username} created successfully!"
                success = True

                # Refresh the user list after adding a new user
                cursor.execute("""
                    SELECT id, username, email, is_admin, created_on
                    FROM users
                    ORDER BY created_on DESC
                """)
                all_users = cursor.fetchall()
                paginator = Paginator(all_users, 5)
                users = paginator.page(1)
            except mysql.connector.Error as err:
                conn.rollback()
                if err.errno == 1062:
                    message = f"Username {username} already exists!"
                else:
                    message = f"Database error: {err}"
                success = False

    cursor.close()
    conn.close()
    print("Users fetched for display:", users.object_list)
    columns = ['id', 'username', 'email', 'is_admin', 'created_on']
    user_dicts = [dict(zip(columns, row)) for row in users.object_list]
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    return render(request, 'create_user.html', {
        'message': message,
        'success': success,
        'users': user_dicts,      # Pass the page object for pagination
        'page_obj': users,   # For compatibility with Django pagination templates
        'user_category': user_category
    })

import json
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods

# method to fetch user cathegory
def get_user_category(username):
    """Fetch user category (admin or regular) from the database"""
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT is_admin FROM users WHERE username = %s", (username,))
    user = cursor.fetchone()
    cursor.close()
    conn.close()
    return user['is_admin'] if user else None

@require_http_methods(["GET"])
def get_user_details(request, user_id):
    print("API get user details view accessed")
    """API endpoint to get user details"""

    try:
        # Connect to database
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Fetch user details
        cursor.execute("""
            SELECT id, username, email, is_admin, created_on 
            FROM users 
            WHERE id = %s
        """, (user_id,))

        user = cursor.fetchone()
        cursor.close()
        conn.close()

        if not user:
            return JsonResponse({'success': False, 'message': 'User not found'})

        return JsonResponse({
            'success': True,
            'user': user
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@require_http_methods(["POST"])
def update_user(request, user_id):
    """API endpoint to update user details"""
    # Check if the user is an admin
    if not request.session.get('is_admin'):
        return JsonResponse({'success': False, 'message': 'You must be an admin to update users'})

    try:
        # Get request data
        data = json.loads(request.body)
        username = data.get('username')
        email = data.get('email')
        is_admin = data.get('is_admin')
        password = data.get('password')

        # Connect to database
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        # Check if username already exists (except for the current user)
        cursor.execute("""
            SELECT id FROM users WHERE username = %s AND id != %s
        """, (username, user_id))

        if cursor.fetchone():
            cursor.close()
            conn.close()
            return JsonResponse({'success': False, 'message': 'Username already exists'})

        # Update user
        if password:
            # Hash the password
            hashed_password = hashlib.sha256(password.encode()).hexdigest()

            # Update user with new password
            cursor.execute("""
                UPDATE users 
                SET username = %s, email = %s, is_admin = %s, password = %s 
                WHERE id = %s
            """, (username, email, is_admin, hashed_password, user_id))
        else:
            # Update user without changing password
            cursor.execute("""
                UPDATE users 
                SET username = %s, email = %s, is_admin = %s 
                WHERE id = %s
            """, (username, email, is_admin, user_id))

        conn.commit()

        # Get updated user data with created_on
        cursor.execute("""
            SELECT created_on FROM users WHERE id = %s
        """, (user_id,))
        user_data = cursor.fetchone()

        cursor.close()
        conn.close()

        return JsonResponse({
            'success': True,
            'message': 'User updated successfully',
            'created_on': user_data['created_on'] if user_data else None
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

from django.shortcuts import render
from django.http import JsonResponse
from datetime import datetime, timedelta

def dashboard(request):
    # check is user is authenticated
    if not request.session.get('is_authenticated'):
        return redirect('login')
    # Stat cards
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    now = datetime.now()
    # Total members
    cursor.execute("SELECT COUNT(*) as total FROM members")
    total_members = cursor.fetchone()['total']
    # Active instructors
    cursor.execute("SELECT COUNT(*) as total FROM instructors WHERE is_active=1")
    active_instructors = cursor.fetchone()['total']
    # Inactive instructors
    cursor.execute("SELECT COUNT(*) as total FROM instructors WHERE is_active=0")
    inactive_instructors = cursor.fetchone()['total']
    # New members this month
    cursor.execute("SELECT COUNT(*) as total FROM members WHERE YEAR(date_of_initiation)=%s AND MONTH(date_of_initiation)=%s", (now.year, now.month))
    new_members_month = cursor.fetchone()['total']
    # New members this year
    cursor.execute("SELECT COUNT(*) as total FROM members WHERE YEAR(date_of_initiation)=%s", (now.year,))
    new_members_year = cursor.fetchone()['total']
    # Growth rate (vs last year)
    cursor.execute("SELECT COUNT(*) as total FROM members WHERE YEAR(date_of_initiation)=%s", (now.year-1,))
    last_year = cursor.fetchone()['total']
    growth_rate = round(((new_members_year - last_year) / last_year * 100), 1) if last_year else 0
    # Recent members
    cursor.execute("SELECT name, date_of_initiation FROM members ORDER BY date_of_initiation DESC LIMIT 7")
    recent_members = cursor.fetchall()
    cursor.close()
    conn.close()
    stats = {
        'total_members': total_members,
        'active_instructors': active_instructors,
        'inactive_instructors': inactive_instructors,
        'new_members_month': new_members_month,
        'new_members_year': new_members_year,
        'growth_rate': growth_rate
    }
    isAdminUser = get_user_category(request.session.get('username'))
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    return render(request, 'dashboard.html', {'stats': stats,
                                              'recent_members': recent_members,
                                              'user_category':user_category})

from django.views.decorators.http import require_GET

@require_GET
def dashboard_metrics_api(request):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    now = datetime.now()
    # Growth over last 24 months
    growth_labels, growth_data = [], []
    for i in range(23, -1, -1):
        dt = (now.replace(day=1) - timedelta(days=30*i))
        cursor.execute("SELECT COUNT(*) as total FROM members WHERE YEAR(date_of_initiation)=%s AND MONTH(date_of_initiation)=%s", (dt.year, dt.month))
        growth_labels.append(dt.strftime('%b %Y'))
        growth_data.append(cursor.fetchone()['total'])

    # Calculate the date 6 months ago
    six_months_ago = (datetime.now() - timedelta(days=180)).date()

    # Members by instructor (last 6 months)
    cursor.execute("""
        SELECT i.name, COUNT(m.id) as total
        FROM instructors i
        LEFT JOIN members m ON m.instructor_id = i.id
            AND m.date_of_initiation >= %s
        GROUP BY i.id
        ORDER BY i.name
    """, (six_months_ago,))
    instructor_labels, instructor_data = [], []
    for row in cursor.fetchall():
        if row['total'] > 0:
            instructor_labels.append(row['name'])
            instructor_data.append(row['total'])
    # Geographic distribution (by state name)
    cursor.execute("""
        SELECT s.name AS state_name, COUNT(*) as total
        FROM members m
        JOIN state s ON m.state = s.id
        GROUP BY s.id
        ORDER BY total DESC
        LIMIT 6
    """)
    geo_labels, geo_data = [], []
    for row in cursor.fetchall():
        geo_labels.append(row['state_name'])
        geo_data.append(row['total'])
    # Top performing instructors
    cursor.execute("""
        SELECT i.name, COUNT(m.id) as total FROM instructors i
        LEFT JOIN members m ON m.instructor_id = i.id
        GROUP BY i.id ORDER BY total DESC LIMIT 5
    """)
    top_labels, top_data = [], []
    for row in cursor.fetchall():
        if row['total'] > 0:
            top_labels.append(row['name'])
            top_data.append(row['total'])
    cursor.close()
    conn.close()
    return JsonResponse({
        'growth': {'labels': growth_labels, 'data': growth_data},
        'instructor': {'labels': instructor_labels, 'data': instructor_data},
        'geo': {'labels': geo_labels, 'data': geo_data},
        'top_instructors': {'labels': top_labels, 'data': top_data}
    })

from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
import json

@require_http_methods(["GET"])
def api_instructor_detail(request, instructor_id):
    print("API instructor detail view accessed")
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT 
            i.*, 
            c.name AS country_name, 
            s.name AS state_name, 
            d.name AS district_name
        FROM instructors i
        LEFT JOIN country c ON i.country = c.id
        LEFT JOIN state s ON i.state = s.id
        LEFT JOIN city d ON i.district = d.id
        WHERE i.id = %s
    """, [instructor_id])
    instructor = cursor.fetchone()
    cursor.close()
    conn.close()
    if not instructor:
        return JsonResponse({'error': 'Instructor not found'}, status=404)
    # Convert date fields to string
    print("Instructor data fetched:", instructor)

    return JsonResponse(instructor)


from django.views.decorators.http import require_http_methods
from django.http import JsonResponse

@require_http_methods(["GET", "POST"])
def api_instructor_update(request, instructor_id):
    print("API update instructor view accessed")
    print("request method:", request.method)
    if request.method == "GET":
        # Fetch instructor details
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM instructors WHERE id = %s", [instructor_id])
        instructor = cursor.fetchone()
        cursor.close()
        conn.close()
        if not instructor:
            return JsonResponse({'error': 'Instructor not found'}, status=404)
        # Convert date fields to string if needed
        if instructor.get('dop'):
            instructor['dop'] = str(instructor['dop'])
        return JsonResponse(instructor)
    elif request.method == "POST":
        try:
            data = json.loads(request.body)
            # validate if name , number (>10 digits) ,age are OK
            if not data.get('name') or not data.get('number') :
                return JsonResponse({'success': False, 'message': 'Name and number are required.'})
            # validate if number is all digits
            if not data.get('number').isdigit():
                return JsonResponse({'success': False, 'message': 'Number must be all digits.'})
            if len(data.get('number', '')) < 10:
                return JsonResponse({'success': False, 'message': 'Number must be at least 10 digits.'})
            if not data.get('age'):
                return JsonResponse({'success': False, 'message': 'Age is required.'})
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE instructors SET
                    name=%s, number=%s, age=%s, gender=%s, associated_since=%s,
                    updeshta_since=%s, address=%s, state=%s, district=%s, country=%s,is_active=%s
                WHERE id=%s
            """, (
                data.get('name'),
                data.get('number'),
                data.get('age') or None,
                data.get('gender') or None,
                data.get('associated_since') or None,
                data.get('updeshta_since') or None,
                data.get('address'),
                data.get('state'),
                data.get('city'),
                data.get('country'),
                data.get('is_active', 1),
                instructor_id
            ))
            conn.commit()
            cursor.close()
            conn.close()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

from django.shortcuts import render, redirect
from django.db import connection

def add_event(request):
    # Fetch instructors for dropdown
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM instructors WHERE is_active=1")
    instructors = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    isAdminUser = get_user_category(request.session.get('username'))
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    message = None
    if request.method == 'POST':
        event_name = request.POST.get('event_name')
        event_date = request.POST.get('event_date')
        instructor_id = request.POST.get('instructor_id')
        coordinator = request.POST.get('event_coordinator')
        location = request.POST.get('event_location')
        country = request.POST.get('event_country')
        state = request.POST.get('event_state')
        district = request.POST.get('event_district')
        event_description = request.POST.get('event_description')

        # Validate inputs
        if not event_name or not event_date or not instructor_id:
            message = "Event name, date, and instructor are required."
        else:
            # Insert event registration
            cursor.execute("SELECT COUNT(*) FROM event_registrations WHERE event_name = %s AND event_date = %s AND location =%s", (event_name, event_date,location))
            if cursor.fetchone()[0] > 0:
                message = "An event with this date, name,and location already exists."
                conn.close()
                return render(request, 'add_event.html', {'instructors': instructors,
                                                          'message': message,
                                                          'user_category': user_category})

        if event_name and event_date and instructor_id:
            cursor.execute("""
                INSERT INTO event_registrations (event_name, event_date, instructor_id,coordinator,location,state, district, country, description)
                VALUES (%s, %s, %s,%s, %s, %s, %s, %s, %s)
            """, [event_name, event_date, instructor_id,coordinator,location,state, district, country, event_description])
            conn.commit()
            message = "Event registered successfully."
            return render(request, 'add_event.html', {'instructors': instructors,
                                                      'message': message,
                                                      'user_category': user_category})
        else:
            message = "All fields are required."
    conn.close()
    return render(request, 'add_event.html', {'instructors': instructors,
                                              'message': message,
                                              'user_category': user_category})

def record_attendance(request):
    # Fetch all events for dropdown
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id, event_name, event_date FROM event_registrations")
    events = [{'id': row[0], 'event_name': row[1], 'event_date': row[2]} for row in cursor.fetchall()]

    selected_event_id = request.GET.get('event_id') or request.POST.get('event_id')
    selected_event = None
    message = None

    # Fetch event details if selected
    if selected_event_id:
        cursor.execute("""
            SELECT 
                e.id, e.event_name, e.event_date, e.instructor_id, i.name,
                e.location,
                country.name AS country_name,
                state.name AS state_name,
                city.name AS city_name,
                country.id AS country_id,
                state.id AS state_id,
                city.id AS city_id,
                e.total_attendance
            FROM event_registrations e
            LEFT JOIN instructors i ON e.instructor_id = i.id
            LEFT JOIN country ON e.country = country.id
            LEFT JOIN state ON e.state = state.id
            LEFT JOIN city ON e.district = city.id
            WHERE e.id = %s
        """, [selected_event_id])
        row = cursor.fetchone()
        if row:
            selected_event = {
                'id': row[0],
                'event_name': row[1],
                'event_date': row[2],
                'instructor_id': row[3],
                'instructor_name': row[4] or '',
                'location': row[5] or '',
                'country': row[6] or '',
                'state': row[7] or '',
                'city': row[8] or '',
                'country_id': row[9] or '',
                'state_id': row[10] or '',
                'city_id': row[11] or '',
                'total_attendance': row[12] or 0
            }

    # Handle attendance register POST
    if request.method == 'POST' and selected_event:
        # Parse dynamic member rows
        members = []
        for key in request.POST:
            if key.startswith('name_'):
                idx = key.split('_')[1]
                try:
                    name = request.POST[f'name_{idx}'].strip()
                    age = request.POST.get(f'age_{idx}', '').strip()
                    contact = request.POST.get(f'contact_{idx}', '').strip()
                    gender = request.POST.get(f'gender_{idx}', '').strip()
                    address = request.POST.get(f'address_{idx}', '').strip()
                    is_new = request.POST.get(f'new_member_{idx}', '') == '1'
                    # validate contact number only digit and len> 10
                    # print error message with name and skip record insertion and continue with next record
                    if not contact.isdigit() or len(contact) < 10:
                        message = f"Invalid contact number for {name}. Must be at least 10 digits long."
                        continue
                    members.append({
                        'name': name, 'age': age, 'contact': contact,
                        'gender': gender, 'address': address, 'is_new': is_new
                    })
                except MultiValueDictKeyError:
                    continue

        # Save new members and count
        new_member_count = 0

        for m in members:
            if m['is_new'] :
                cursor.execute("""
                    INSERT INTO members (name, number, age, gender, address, state, district, country, instructor_id, date_of_initiation)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, [
                    m['name'], m['contact'], m['age'], m['gender'], m['address'],
                    selected_event['state_id'], selected_event['city_id'], selected_event['country_id'],
                    selected_event['instructor_id'], selected_event['event_date']
                ])
                new_member_count += 1

        # Update total_attendance
        cursor.execute("""
            UPDATE event_registrations
            SET total_attendance = total_attendance + %s
            WHERE id = %s
        """, [len(members), selected_event['id']])

        # record the attendance in event_attendance table
        for m in members:
            cursor.execute("""
                INSERT INTO event_attendance (
                    event_id,
                    member_name,
                    age,
                    contact_number,
                    gender,
                    address,
                    is_new_member
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, [
                selected_event['id'],
                m['name'],
                m.get('age'),
                m.get('contact'),
                m.get('gender'),
                m.get('address'),
                1 if m.get('is_new') else 0
            ])
        conn.commit()
        conn.close()
        message = f"Attendance recorded. {new_member_count} new member(s) registered successfully."
        # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    return render(request, 'record_attendance.html', {
        'events': events,
        'selected_event': selected_event,
        'selected_event_id': str(selected_event_id) if selected_event_id else '',
        'message': message,
        'user_category': user_category
    })

# views.py
from django.http import JsonResponse
from django.db import connection

def get_countries(request):
    print("API get countries view accessed")
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT id, name FROM country ORDER BY name")
    countries = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    conn.close()
    return JsonResponse({'countries': countries})

def get_states(request):
    conn = get_db_connection()
    cursor = conn.cursor()
    country_id = request.GET.get('country_id')
    if not country_id:
        return JsonResponse({'states': []})

    cursor.execute("SELECT id, name FROM state WHERE country_id = %s ORDER BY name", [country_id])
    states = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    conn.close()
    return JsonResponse({'states': states})

def get_cities(request):
    conn = get_db_connection()
    cursor = conn.cursor()
    state_id = request.GET.get('state_id')
    if not state_id:
        return JsonResponse({'cities': []})

    cursor.execute("SELECT id, name FROM city WHERE state_id = %s ORDER BY name", [state_id])
    cities = [{'id': row[0], 'name': row[1]} for row in cursor.fetchall()]
    conn.close()

    return JsonResponse({'cities': cities})

def session_timeout(request):
    return render(request, 'session_timeout.html')

def get_db_conn():
    return mysql.connector.connect(
        host=settings.DB_HOST,
        user=settings.DB_USER,
        password=settings.DB_PASSWORD,
        database=settings.DB_NAME
    )

def forgot_password(request):
    message = None
    message_type = "error"
    if request.method == "POST":
        email = request.POST.get("email")
        conn = get_db_conn()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id, username FROM users WHERE email=%s", (email,))
        user = cur.fetchone()
        if user:
            # Generate token
            token = secrets.token_urlsafe(32)
            RESET_TOKENS[token] = {
                "user_id": user["id"],
                "username": user["username"],
                "expires": datetime.datetime.now() + datetime.timedelta(hours=1)
            }
            # Simulate sending email (show link on page for demo)
            reset_link = request.build_absolute_uri(reverse('reset_password', args=[token]))
            message = f"Reset link sent!<br>Your username: <b>{user['username']}</b><br><a href='{reset_link}'>Click here to reset your password</a> (valid for 1 hour)."
            message_type = "success"
        else:
            message = "No user found with that email address."
        cur.close()
        conn.close()

    return render(request, "forgot_password.html", {"message": message, "message_type": message_type})

def reset_password(request, token):
    info = RESET_TOKENS.get(token)
    message = None
    message_type = "error"
    username = None
    if not info or info["expires"] < datetime.datetime.now():
        message = "Invalid or expired reset link."
        return render(request, "reset_password.html", {"message": message, "message_type": message_type})
    username = info["username"]
    if request.method == "POST":
        pwd = request.POST.get("password")
        cpwd = request.POST.get("confirm_password")
        if pwd != cpwd:
            message = "Passwords do not match."
        elif len(pwd) < 8:
            message = "Password must be at least 8 characters."
        else:
            # Hash password (simple SHA256 for demo; use bcrypt/argon2 in production)
            hashed = hashlib.sha256(pwd.encode()).hexdigest()
            conn = get_db_conn()
            cur = conn.cursor()
            cur.execute("UPDATE users SET password=%s WHERE id=%s", (hashed, info["user_id"]))
            conn.commit()
            cur.close()
            conn.close()
            RESET_TOKENS.pop(token, None)
            message = "Password reset successful! You can now <a href='%s'>login</a>." % reverse('login')
            message_type = "success"
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    return render(request, "reset_password.html", {
        "message": message,
        "message_type": message_type,
        "username": username,
        "user_category": user_category
    })

def view_events(request):
    print("View events page accessed")
    # Fetch all events for filter dropdown
    conn = get_db_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, event_name, event_date FROM event_registrations ORDER BY event_date DESC")
    events = cur.fetchall()
    #[dict(zip([col[0] for col in cur.description], row)) for row in cur.fetchall()]
    print("Events fetched for display:", events)
    selected_event_id = request.GET.get('event_id')
    page_number = request.GET.get('page', 1)
    print("Selected event ID:", selected_event_id, "Page number:", page_number)
    # Event summary
    event_summary = None
    if selected_event_id:
        cur.execute("""
            SELECT er.*, i.name as instructor_name
            FROM event_registrations er
            LEFT JOIN instructors i ON er.instructor_id = i.id
            WHERE er.id = %s
        """, [selected_event_id])
        event_summary = cur.fetchone()
        #if row:
            #event_summary = dict(zip([col[0] for col in cur.description], row))
        print("Event summary fetched:", event_summary)
    # Attendance data
    if selected_event_id:
        cur.execute("""
            SELECT * FROM event_attendance
            WHERE event_id = %s
            ORDER BY attended_on DESC
        """, [selected_event_id])
    else:
        cur.execute("""
            SELECT * FROM event_attendance
            ORDER BY attended_on DESC
        """)
    attendance =  cur.fetchall()

    paginator = Paginator(attendance, 10)
    attendance_page = paginator.get_page(page_number)
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    return render(request, 'view_events.html', {
        'events': events,
        'selected_event_id': selected_event_id or '',
        'event_summary': event_summary,
        'attendance_page': attendance_page,
        'user_category': user_category
    })

# Excel download view
def download_event_attendance(request):
    event_id = request.GET.get('event_id')
    # Fetch event summary
    event_summary = {}
    conn = get_db_conn()
    cur = conn.cursor(dictionary=True)
    if event_id:
        cur.execute("""
            SELECT er.*, i.name as instructor_name,
                   c.name as country_name,
                   s.name as state_name,
                   d.name as city_name
            FROM event_registrations er
            LEFT JOIN instructors i ON er.instructor_id = i.id
            LEFT JOIN country c ON er.country = c.id
            LEFT JOIN state s ON er.state = s.id
            LEFT JOIN city d ON er.district = d.id
            WHERE er.id = %s
        """, [event_id])
        event_summary = cur.fetchone()
        print("Event summary fetched for Excel:", event_summary)
    # Fetch attendance

    if event_id:
        cur.execute("""
            SELECT member_name, age, contact_number, gender, address, attended_on, is_new_member
            FROM event_attendance WHERE event_id = %s
            ORDER BY attended_on DESC
        """, [event_id])
    else:
        cur.execute("""
            SELECT member_name, age, contact_number, gender, address, attended_on, is_new_member
            FROM event_attendance
            ORDER BY attended_on DESC
        """)
    attendance = cur.fetchall()



    # Create Excel
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Event Summary"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    header_fill = PatternFill("solid", fgColor="2563eb")
    data_font = Font(name="Calibri", size=11)
    data_fill = PatternFill("solid", fgColor="f3f6fa")
    border = Border(
        left=Side(style="thin", color="B0B7C3"),
        right=Side(style="thin", color="B0B7C3"),
        top=Side(style="thin", color="B0B7C3"),
        bottom=Side(style="thin", color="B0B7C3"),
    )
    align = Alignment(vertical="center", horizontal="left", wrap_text=True)

    # Write event summary
    if event_summary:
        ws1.append(["Field", "Value"])
        # Map DB keys to user-friendly labels
        field_labels = {
            "event_name": "Event Name",
            "event_date": "Event Date",
            "instructor_name": "Instructor",
            "coordinator": "Coordinator",
            "location": "Location",
            "country_name": "Country",
            "state_name": "State",
            "city_name": "City",
            "total_attendance": "Total Attendance",
            "description": "Description",
            # Add more as needed
        }

        for key in ["event_name", "event_date", "instructor_name", "coordinator", "location",
                    "country_name", "state_name", "city_name", "total_attendance", "description"]:
            if key in event_summary:
                label = field_labels.get(key, key.replace("_", " ").title())
                value = event_summary[key] if event_summary[key] is not None else ""
                ws1.append([label, str(value)])
        # Style header
        for col in range(1, 3):
            cell = ws1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align
            cell.border = border
        # Style data
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.font = data_font
                cell.fill = data_fill
                cell.alignment = align
                cell.border = border
    else:
        ws1["A1"] = "No event selected"
        ws1["A1"].font = header_font
        ws1["A1"].fill = header_fill
        ws1["A1"].alignment = align
        ws1["A1"].border = border

    # Attendance sheet
    ws2 = wb.create_sheet("Attendance")
    headers = ["Name", "Age", "Contact", "Gender", "Address", "Attended On", "New Member?"]
    ws2.append(headers)
    # Style header
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align
        cell.border = border

    # Data rows
    for row in attendance:
        ws2.append([
            row['member_name'],
            row['age'],
            row['contact_number'],
            row['gender'],
            row['address'],
            row['attended_on'].strftime("%Y-%m-%d %H:%M") if row['attended_on'] else "",
            "Yes" if row['is_new_member'] else "No"
        ])
    for r in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=len(headers)):
        for cell in r:
            cell.font = data_font
            cell.fill = data_fill
            cell.alignment = align
            cell.border = border

    # Auto column width
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 4

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "event_attendance.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@require_http_methods(["GET", "POST"])
def upload_attendance(request):
    # Fetch all events for dropdown
    conn = get_db_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, event_name, event_date FROM event_registrations ORDER BY event_date DESC")
    events = cur.fetchall()
    cur.close()
    conn.close()
    # fetch user category
    isAdminUser = get_user_category(request.session['username'])
    print("User category fetched admin status:", isAdminUser)
    if isAdminUser == True:
        user_category = 'admin'
    else:
        user_category = 'standard'
    if request.method == "POST":
        event_id = request.POST.get('event_id')
        file = request.FILES.get('attendance_file')
        if not event_id or not file:
            messages.error(request, "Please select an event and upload a file.")
            return render(request, 'upload_attendance.html', {'events': events, 'user_category': user_category})

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            expected = ['Name', 'Age', 'Contact', 'Gender', 'Address', 'Attended On', 'New Member?']
            if headers != expected:
                messages.error(request, "Invalid template. Please use the provided template.")
                return render(request, 'upload_attendance.html', {'events': events, 'user_category': user_category})

            conn = get_db_connection()
            cur = conn.cursor()
            count = 0
            new_member_count = 0
            total_members = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    cur.execute("""
                        INSERT INTO event_attendance
                        (event_id, member_name, age, contact_number, gender, address, attended_on, is_new_member)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        event_id,
                        row[0],
                        int(row[1]) if row[1] else None,
                        row[2] or "",
                        row[3] or "",
                        row[4] or "",
                        row[5] if row[5] else None,
                        1 if (str(row[6]).strip().lower() == "yes") else 0
                    ))
                    count += 1

                    # Register as new member for rows marked as yes  if needed or an old member who doesn't exist in database .
                    if str(row[6]).strip().lower() == "yes":
                        # Fetch event location info
                        cur.execute("""
                            SELECT country, state, district,instructor_id FROM event_registrations WHERE id = %s
                        """, (event_id,))
                        event_loc = cur.fetchone()
                        country = event_loc[0] if event_loc else None
                        state = event_loc[1] if event_loc else None
                        district = event_loc[2] if event_loc else None
                        instructor_id = event_loc[3] if event_loc else None


                        # Insert into member table
                        cur.execute("""
                            INSERT INTO members
                            (name, age, gender, address, state, district, country, event_id, date_of_initiation, number,instructor_id)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)
                        """, (
                            row[0],  # name
                            str(row[1]) if row[1] else None,  # age as string
                            row[3] or "",  # gender
                            row[4] or "",  # address
                            state,
                            district,
                            country,
                            event_id,
                            row[5] if row[5] else None,  # attended_on as date_of_initiation
                            row[2] or "" , # contact number
                            instructor_id
                        ))
                        new_member_count += 1
            # Update total attendance for the event

            cur.execute("""
                UPDATE event_registrations
                SET total_attendance = %s
                WHERE id = %s
            """, (count, event_id))

            conn.commit()
            cur.close()
            conn.close()
            messages.success(request, f"Successfully uploaded {count} attendance records. {new_member_count} new member(s) registered.")
            return redirect(reverse('upload_attendance'))
        except Exception as e:
            print("Error processing file:", e)
            messages.error(request, f"Error processing file: {e}")
    return render(request, 'upload_attendance.html', {'events': events})

def isexisting_member(name, contact):
    conn = get_db_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT COUNT(*) as total FROM members WHERE name=%s AND number=%s", (name, contact))
    exists = cur.fetchone()['total'] > 0
    cur.close()
    conn.close()
    return exists

def ajax_events(request):
        # Get filters
        print("AJAX events view accessed")
        name = request.GET.get('name', '').strip()
        coordinator = request.GET.get('coordinator', '').strip()
        date = request.GET.get('date', '').strip()
        instructor = request.GET.get('instructor', '').strip()

        conn = get_db_conn()
        cur = conn.cursor(dictionary=True)
        sql = """
            SELECT
                e.id,
                e.event_name,
                e.event_date,
                e.coordinator,
                e.location,
                country.name AS country,
                state.name AS state,
                city.name AS district,
                e.total_attendance,
                e.description,
                e.instructor_id,
                i.name AS instructor_name
            FROM event_registrations e
            LEFT JOIN instructors i ON e.instructor_id = i.id
            LEFT JOIN country ON e.country = country.id
            LEFT JOIN state ON e.state = state.id
            LEFT JOIN city ON e.district = city.id
            WHERE 1=1
        """
        params = []
        if name:
            sql += " AND e.event_name LIKE %s"
            params.append(f"%{name}%")
        if coordinator:
            sql += " AND e.coordinator LIKE %s"
            params.append(f"%{coordinator}%")
        if date:
            sql += " AND e.event_date = %s"
            params.append(date)
        if instructor:
            sql += " AND e.instructor_id = %s"
            params.append(instructor)
        sql += " ORDER BY e.id DESC LIMIT 100"
        cur.execute(sql, params)
        events = []
        for row in cur.fetchall():
            events.append({
                'id': row['id'],
                'name': row['event_name'],
                'date': row['event_date'].strftime('%Y-%m-%d') if row['event_date'] else '',
                'coordinator': row['coordinator'],
                'location': row['location'],
                'state': row['state'],
                'district': row['district'],
                'country': row['country'],
                'total_attendance': row['total_attendance'],
                'description': row['description'],
                'instructor_id': row['instructor_id'],
                'instructor': row['instructor_name'] if row['instructor_name'] else 'Unspecified'
            })
        cur.close()
        conn.close()
        return JsonResponse({'events': events})

@csrf_exempt
def ajax_events_edit(request):
    print("AJAX events edit view accessed ")
    if request.method == 'POST':
        data = json.loads(request.body)
        print("Data received for edit:", data)
        conn = get_db_conn()
        cur = conn.cursor()
        sql = """
            UPDATE event_registrations
            SET event_name=%s,
                event_date=%s,
                coordinator=%s,
                location=%s,
                instructor_id=%s,
                country=%s,
                state=%s,
                district=%s,
                description=%s,
                total_attendance=%s
            WHERE id=%s
        """
        cur.execute(sql, (
            data['name'],
            data['date'],
            data['coordinator'],
            data['location'],
            data['instructor_id'],
            data['country'],
            data['state'],
            data['district'],
            data.get('description', ''),
            data.get('total_attendance', 0),
            data['id']
        ))
        conn.commit()
        cur.close()
        conn.close()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False})

def ajax_events_download(request):
    name = request.GET.get('name', '').strip()
    coordinator = request.GET.get('coordinator', '').strip()
    date = request.GET.get('date', '').strip()
    instructor = request.GET.get('instructor', '').strip()

    conn = get_db_conn()
    cur = conn.cursor(dictionary=True)
    sql = """
        SELECT e.event_name, e.event_date, e.coordinator, e.location, e.state, e.district, e.country,
               e.total_attendance, e.description, i.name AS instructor_name
        FROM event_registrations e
        LEFT JOIN instructors i ON e.instructor_id = i.id
        WHERE 1=1
    """
    params = []
    if name:
        sql += " AND e.event_name LIKE %s"
        params.append(f"%{name}%")
    if coordinator:
        sql += " AND e.coordinator LIKE %s"
        params.append(f"%{coordinator}%")
    if date:
        sql += " AND e.event_date = %s"
        params.append(date)
    if instructor:
        sql += " AND e.instructor_id = %s"
        params.append(instructor)
    sql += " ORDER BY e.id DESC"
    cur.execute(sql, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Events"
    headers = [
        'Event Name', 'Date', 'Coordinator', 'Location', 'State', 'District', 'Country',
        'Total Attendance', 'Description', 'Instructor'
    ]
    ws.append(headers)

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
    header_fill = PatternFill("solid", fgColor="2563eb")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin', color='CCCCCC'),
                         right=Side(style='thin', color='CCCCCC'),
                         top=Side(style='thin', color='CCCCCC'),
                         bottom=Side(style='thin', color='CCCCCC'))
    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows styling
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for r in rows:
        ws.append([
            r['event_name'],
            r['event_date'].strftime('%Y-%m-%d') if r['event_date'] else '',
            r['coordinator'],
            r['location'],
            r['state'],
            r['district'],
            r['country'],
            r['total_attendance'],
            r['description'],
            r['instructor_name'] or ''
        ])
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value)
                if len(val) > max_length:
                    max_length = len(val)
            except:
                pass
        ws.column_dimensions[col_letter].width = max(14, min(max_length + 2, 40))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Events.xlsx'
    wb.save(response)
    return response

def get_events_for_dropdown():
    conn = get_db_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT id, event_name FROM event_registrations ORDER BY event_name")
    events = cur.fetchall()
    cur.close()
    conn.close()
    return events

@csrf_exempt
def public_register(request):
    print("Public registration page accessed")
    message = ''
    message_type = ''
    year = datetime.now().year
    events = get_events_for_dropdown()

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        number = request.POST.get('number', '').strip()
        email = request.POST.get('email', '').strip()
        age = request.POST.get('age', '').strip()
        gender = request.POST.get('gender', '').strip()
        event_id = request.POST.get('event', '').strip()
        company = request.POST.get('company', '').strip()
        address = request.POST.get('address', '').strip()
        member_country = request.POST.get('member_country', '').strip()
        member_state = request.POST.get('member_state', '').strip()
        member_city = request.POST.get('member_city', '').strip()
        feedback = request.POST.get('feedback', '').strip()
        instructor_id = request.POST.get('instructor', '').strip()
        date_of_initiation = request.POST.get('date_of_initiation', '').strip()
        # validate name as non digit
        if any(char.isdigit() for char in name):
            form_data = request.POST.copy()
            message = 'Name must contain only letters.'
            message_type = 'error'
            instructors = get_instructors()
            events = get_events()
            return render(request, 'public_register.html', {
                'message': message,
                'message_type': message_type,
                'form_data': form_data,
                'instructors': instructors,
                'events': events
            })
        if not date_of_initiation:
            date_of_initiation = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if email:
            # validate the form fields
            form_data = request.POST.copy()
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_regex, email):
                instructors = get_instructors()
                events = get_events()
                return render(request, 'public_register.html', {
                    'message': 'Invalid email format.',
                    'message_type': 'error',
                    'form_data': form_data,
                    'instructors': instructors,
                    'events': events
                })
        if not name or not number or not age :
            message = 'Name, Contact Number, And Age are required.'
            message_type = 'error'
        else:
            conn = get_db_conn()
            cur = conn.cursor()
            cur.execute("""
                            INSERT INTO members
                            (name, number, email, address,age,gender, country, state,district,company, notes, instructor_id, date_of_initiation,event_id)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s,%s, %s, %s)
                        """, (
                name, number, email, address, age, gender, member_country, member_state, member_city, company, feedback, instructor_id,
                date_of_initiation, event_id
            ))
            conn.commit()
            cur.close()
            conn.close()
            message = 'Registration successful!'
            message_type = 'success'

    def get_public_instructors():
        conn = mysql.connector.connect(
            host=settings.DB_HOST,
            user=settings.DB_USER,
            password=settings.DB_PASSWORD,
            database=settings.DB_NAME
        )
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT id, name FROM instructors")
        instructors = cur.fetchall()
        cur.close()
        conn.close()
        return instructors

    instructors = get_public_instructors()
    return render(request, 'public_register.html', {
        'message': message,
        'message_type': message_type,
        'events': events,
        'instructors': instructors,
        'year': year
    })

# views.py (backend endpoint)
from django.http import JsonResponse
from django.views.decorators.http import require_GET

@require_GET
def ajax_eventsbyDate(request):
    date = request.GET.get('date', '').strip()
    conn = get_db_conn()
    cur = conn.cursor(dictionary=True)
    sql = """
        SELECT id, event_name, event_date
        FROM event_registrations
        WHERE 1=1
    """
    params = []
    if date:
        sql += " AND event_date = %s"
        params.append(date)
    sql += " ORDER BY event_date DESC LIMIT 100"
    cur.execute(sql, params)
    events = []
    for row in cur.fetchall():
        events.append({
            'id': row['id'],
            'name': row['event_name'],
            'date': row['event_date'].strftime('%Y-%m-%d') if row['event_date'] else ''
        })
    cur.close()
    conn.close()
    return JsonResponse({'events': events})

def thank_you(request):
    return render(request, 'thank_you.html')
